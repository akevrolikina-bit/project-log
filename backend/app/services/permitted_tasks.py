"""Load the permitted-task list from the Issues CHANGE Excel export.

The "LOG" sheet has two kinds of rules:

1. **Key-based rules** (rows with a Jira key in column D):
       A=Project, B=Type, C=Group, D=Key, E=Title, F="да"/"нет", G=Exclusions

2. **Project+Type rules** (rows with Project+Type but no Key):
       A=Project, B=Type, F="да"/"нет", G=Exclusions
       These act as catch-all rules for any key under that project/type combo.

Exclusion formats in column G:
    - "Носов Михаил"                  — this person is excluded
    - "все, кроме Носова Михаила"     — only the named person is allowed
    - empty                           — no exclusions

Invest columns (I/J):
    - I = invest direction (concrete name like ``MENA`` / ``Alphyn``, or
      ambiguous ``MENA / другой``)
    - J = allocation method; empty J means auto **only** when I is concrete
"""

from __future__ import annotations

import io
import os
import re
from collections.abc import Iterable
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

from app.config import settings

_DEFAULT_PATH = Path(settings.permitted_tasks_path)


COMMENT_RULE_STRICT = "strict"
COMMENT_RULE_LENIENT = "lenient"

INVEST_AUTO = "auto"
INVEST_MANUAL_PERCENT = "manual_percent"
INVEST_MANUAL_PROJECT = "manual_project"
INVEST_BUH_COMPANY = "buh_company"
INVEST_KEYWORD = "keyword"
INVEST_PLAN_FTE = "plan_fte"

# Shown first in dropdowns and reports; remaining names stay alphabetical.
_PREFERRED_INVEST_ORDER = ("MENA",)


def sort_invest_projects(projects: Iterable[str]) -> list[str]:
    """Return unique project names with MENA first, then A–Z."""
    unique = {p.strip() for p in projects if p and str(p).strip()}
    preferred = [p for p in _PREFERRED_INVEST_ORDER if p in unique]
    preferred_set = set(preferred)
    rest = sorted(p for p in unique if p not in preferred_set)
    return preferred + rest


@dataclass
class KeyRule:
    """Rule for a specific Jira key."""

    project: str
    task_type: str
    key: str
    title: str
    permitted: bool
    excluded_users: list[str] = field(default_factory=list)
    only_users: list[str] = field(default_factory=list)
    comment_rule: str = ""
    invest_direction: str = ""
    invest_allocation: str = ""
    keyword_rules: dict[str, list[str]] = field(default_factory=dict)


@dataclass
class ProjectTypeRule:
    """Catch-all rule for a Project + Type combination."""

    project: str
    task_type: str
    permitted: bool
    excluded_users: list[str] = field(default_factory=list)
    only_users: list[str] = field(default_factory=list)
    comment_rule: str = ""
    invest_direction: str = ""
    invest_allocation: str = ""
    keyword_rules: dict[str, list[str]] = field(default_factory=dict)


@dataclass
class CheckVerdict:
    """Result of checking a single worklog entry against the registry."""

    permitted: bool
    reason: str
    rule_type: str  # "key", "project_type", or "unknown"


def _parse_exclusions(raw: str) -> tuple[list[str], list[str]]:
    """Parse the exclusion column.

    Returns (excluded_users, only_users).
    - excluded_users: people who are NOT allowed even though the task is "да".
    - only_users: if non-empty, ONLY these people are allowed.
    """
    raw = raw.strip()
    if not raw:
        return [], []

    krome_match = re.match(
        r"все[,.]?\s*кроме\s+(.+)", raw, re.IGNORECASE
    )
    if krome_match:
        names_part = krome_match.group(1).strip()
        only = [n.strip() for n in re.split(r"[,;]", names_part) if n.strip()]
        return [], only

    excluded = [n.strip() for n in re.split(r"[,;]", raw) if n.strip()]
    return excluded, []


def _name_key(name: str) -> str:
    """Normalize a Russian name for comparison.

    Strips case-endings by taking the first 3 chars of surname + first 3 chars
    of first name, so "Носов Михаил" and "Носова Михаила" both produce "нос мих".
    """
    parts = name.strip().split()
    return " ".join(p[:3].lower() for p in parts if p)


def _names_match(a: str, b: str) -> bool:
    return _name_key(a) == _name_key(b)


class PermittedTasksRegistry:
    """Two-level lookup: by Jira key, then by project+type."""

    def __init__(self) -> None:
        self._key_rules: dict[str, KeyRule] = {}
        self._pt_rules: dict[tuple[str, str], ProjectTypeRule] = {}
        self.time_limited_keys: set[str] = set()
        self.general_rules: list[str] = []

    def add_key_rule(self, rule: KeyRule) -> None:
        self._key_rules[rule.key] = rule

    def add_project_type_rule(self, rule: ProjectTypeRule) -> None:
        self._pt_rules[(rule.project, rule.task_type)] = rule

    def check(self, key: str, project: str, task_type: str, username: str) -> CheckVerdict:
        """Check whether this worklog entry is permitted for the given user.

        Lookup order:
        1. Specific key rule
        2. Project + Type catch-all rule
        3. Unknown (no rule found)
        """
        kr = self._key_rules.get(key)
        if kr is not None:
            return self._apply_rule(
                permitted=kr.permitted,
                excluded_users=kr.excluded_users,
                only_users=kr.only_users,
                username=username,
                rule_type="key",
                label=kr.key,
            )

        ptr = self._pt_rules.get((project, task_type))
        if ptr is not None:
            return self._apply_rule(
                permitted=ptr.permitted,
                excluded_users=ptr.excluded_users,
                only_users=ptr.only_users,
                username=username,
                rule_type="project_type",
                label=f"{project}/{task_type}",
            )

        return CheckVerdict(
            permitted=False,
            reason=f"Задача {key} ({project}/{task_type}) отсутствует в списке разрешённых",
            rule_type="unknown",
        )

    @staticmethod
    def _apply_rule(
        *,
        permitted: bool,
        excluded_users: list[str],
        only_users: list[str],
        username: str,
        rule_type: str,
        label: str,
    ) -> CheckVerdict:
        if not permitted:
            return CheckVerdict(
                permitted=False,
                reason=f"Задача {label} запрещена для списания",
                rule_type=rule_type,
            )

        if only_users:
            if not any(_names_match(username, name) for name in only_users):
                return CheckVerdict(
                    permitted=False,
                    reason=f"Задача {label} разрешена только для: {', '.join(only_users)}",
                    rule_type=rule_type,
                )

        if excluded_users:
            for excl_name in excluded_users:
                if _names_match(username, excl_name):
                    return CheckVerdict(
                        permitted=False,
                        reason=f"Сотруднику {username} запрещено списывать на {label}",
                        rule_type=rule_type,
                    )

        return CheckVerdict(permitted=True, reason="OK", rule_type=rule_type)

    def get_invest_info(
        self, key: str, project: str, task_type: str
    ) -> tuple[str, str] | None:
        """Return (invest_direction, invest_allocation) or None.

        Lookup order mirrors ``check()``: key rule first, then project+type.
        Returns None when the matched rule has no usable invest data
        (missing direction, missing allocation method, or both).
        """
        kr = self._key_rules.get(key)
        if kr is not None and kr.invest_direction and kr.invest_allocation:
            return kr.invest_direction, kr.invest_allocation

        ptr = self._pt_rules.get((project, task_type))
        if ptr is not None and ptr.invest_direction and ptr.invest_allocation:
            return ptr.invest_direction, ptr.invest_allocation

        return None

    def resolve_keyword_invest(
        self, project: str, task_type: str, title: str
    ) -> str | None:
        """Resolve invest project from task title using keyword rules."""
        ptr = self._pt_rules.get((project, task_type))
        if ptr is None or ptr.invest_allocation != INVEST_KEYWORD:
            return None
        return match_keyword_project(title, ptr.keyword_rules)

    def list_concrete_invest_directions(self) -> list[str]:
        """Return unique invest project names from rules (e.g. MENA, Alphyn).

        MENA is listed first; remaining names are alphabetical.
        Ambiguous labels like ``MENA / другой`` are excluded.
        """
        directions: set[str] = set()
        for rule in self._key_rules.values():
            if rule.invest_allocation and _is_concrete_invest_direction(
                rule.invest_direction
            ):
                directions.add(rule.invest_direction.strip())
        for rule in self._pt_rules.values():
            if rule.invest_allocation and _is_concrete_invest_direction(
                rule.invest_direction
            ):
                directions.add(rule.invest_direction.strip())
        return sort_invest_projects(directions)

    def get_comment_rule(self, key: str, project: str, task_type: str) -> str:
        """Return the comment rule for a task: 'strict', 'lenient', or ''."""
        kr = self._key_rules.get(key)
        if kr is not None:
            return kr.comment_rule

        ptr = self._pt_rules.get((project, task_type))
        if ptr is not None:
            return ptr.comment_rule

        return ""

    @property
    def key_rules(self) -> dict[str, KeyRule]:
        return dict(self._key_rules)

    @property
    def project_type_rules(self) -> dict[tuple[str, str], ProjectTypeRule]:
        return dict(self._pt_rules)


_TIME_LIMITED_KEYS = {"BUH-72900", "BUH-115258"}


_INVEST_ALLOC_MAP = {
    "руками, задавая процент": INVEST_MANUAL_PERCENT,
    "руками": INVEST_MANUAL_PROJECT,
    "по buh company из дополнительной выгрузки": INVEST_BUH_COMPANY,
}


def _parse_keyword_rules(raw: str) -> dict[str, list[str]]:
    """Parse multi-line keyword rules from column J."""
    rules: dict[str, list[str]] = {}
    for line in raw.splitlines():
        line = line.strip()
        if not line or line.lower().startswith("по ключевым"):
            continue
        if "=" not in line:
            continue
        project, _, keywords_part = line.partition("=")
        project = project.strip()
        keywords = [k.strip() for k in keywords_part.split(",") if k.strip()]
        if project and keywords:
            rules[project] = keywords
    return rules


def match_keyword_project(title: str, keyword_rules: dict[str, list[str]]) -> str | None:
    """Return the first invest project whose keyword appears in *title*."""
    title_lower = title.lower()
    for project, keywords in keyword_rules.items():
        for keyword in keywords:
            if keyword.lower() in title_lower:
                return project
    return None


def _is_concrete_invest_direction(direction: str) -> bool:
    """True when column I names a single invest project (e.g. MENA, Alphyn).

    Values like ``MENA / другой`` are ambiguous and need an allocation method
    in column J; they must not be treated as automatic.
    """
    direction = direction.strip()
    return bool(direction) and "/" not in direction


def _normalize_invest_allocation(raw: str, invest_direction: str) -> str:
    """Map the raw invest allocation text to one of the INVEST_* constants."""
    raw_stripped = raw.strip()
    raw_lower = raw_stripped.lower()
    if not raw_lower:
        # Empty J → auto only for a concrete project name in I.
        return INVEST_AUTO if _is_concrete_invest_direction(invest_direction) else ""
    if raw_lower.startswith("по ключевым словам"):
        return INVEST_KEYWORD
    if "пропорции от задаваемого плана" in raw_lower:
        return INVEST_PLAN_FTE
    return _INVEST_ALLOC_MAP.get(raw_lower, raw_stripped)


def _parse_comment_rule(raw: str) -> str:
    """Classify the comment rule text into a machine-usable constant."""
    raw = raw.strip().lower()
    if not raw:
        return ""
    if "может не быть" in raw or "может быть комментарий" in raw:
        return COMMENT_RULE_LENIENT
    if "не может быть пустым" in raw:
        return COMMENT_RULE_STRICT
    return COMMENT_RULE_STRICT


def load_permitted_tasks(path: str | Path | None = None) -> PermittedTasksRegistry:
    """Parse the Excel file and return a registry.

    Parameters
    ----------
    path : path to the .xlsx file.  Defaults to
           ``data/input/Issues CHANGE (3).xlsx`` relative to project root.
           Can also be set via ``PERMITTED_TASKS_PATH`` env var.
    """
    if path is None:
        path = os.environ.get("PERMITTED_TASKS_PATH", str(_DEFAULT_PATH))
    path = Path(path)

    if not path.exists():
        raise FileNotFoundError(f"Permitted-tasks file not found: {path}")

    try:
        file_bytes = path.read_bytes()
    except PermissionError as exc:
        raise PermissionError(
            f"Permitted-tasks file is locked (close it in Excel): {path}"
        ) from exc

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)

    sheet_name = "LOG"
    if sheet_name not in wb.sheetnames:
        sheet_name = wb.sheetnames[0]

    ws = wb[sheet_name]
    registry = PermittedTasksRegistry()
    registry.time_limited_keys = set(_TIME_LIMITED_KEYS)

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:
            continue

        proj = str(row[0]).strip() if row[0] else ""
        typ = str(row[1]).strip() if row[1] else ""
        group = str(row[2]).strip() if row[2] else ""
        key = str(row[3]).strip() if row[3] else ""
        title = str(row[4]).strip() if row[4] else ""
        status_raw = str(row[5]).strip().lower() if row[5] else ""
        excl_raw = str(row[6]).strip() if len(row) > 6 and row[6] else ""
        comment_rule_raw = str(row[7]).strip() if len(row) > 7 and row[7] else ""
        invest_dir_raw = str(row[8]).strip() if len(row) > 8 and row[8] else ""
        invest_alloc_raw = str(row[9]).strip() if len(row) > 9 and row[9] else ""

        if status_raw not in ("да", "нет"):
            if proj and not key and not typ:
                text = proj.strip()
                if text and not text.startswith("Project"):
                    registry.general_rules.append(text)
            continue

        permitted = status_raw == "да"
        excluded_users, only_users = _parse_exclusions(excl_raw)
        comment_rule = _parse_comment_rule(comment_rule_raw)
        invest_allocation = _normalize_invest_allocation(
            invest_alloc_raw, invest_dir_raw
        )
        keyword_rules = (
            _parse_keyword_rules(invest_alloc_raw)
            if invest_allocation == INVEST_KEYWORD
            else {}
        )
        # Drop invest metadata when allocation could not be determined
        # (e.g. I="MENA / другой" with empty J — ambiguous, not auto).
        invest_direction = invest_dir_raw if invest_allocation else ""

        if key and "-" in key:
            registry.add_key_rule(
                KeyRule(
                    project=proj,
                    task_type=typ,
                    key=key,
                    title=title,
                    permitted=permitted,
                    excluded_users=excluded_users,
                    only_users=only_users,
                    comment_rule=comment_rule,
                    invest_direction=invest_direction,
                    invest_allocation=invest_allocation,
                    keyword_rules=keyword_rules,
                )
            )
        elif proj and typ:
            registry.add_project_type_rule(
                ProjectTypeRule(
                    project=proj,
                    task_type=typ,
                    permitted=permitted,
                    excluded_users=excluded_users,
                    only_users=only_users,
                    comment_rule=comment_rule,
                    invest_direction=invest_direction,
                    invest_allocation=invest_allocation,
                    keyword_rules=keyword_rules,
                )
            )

    wb.close()
    return registry
