"""Generate an Excel (.xlsx) report.

Sheets: Сводка, Распределение, Инвест-направления, Простои GENERAL-122,
Ошибки, Недобор часов.
Styled according to docs/brandbook.html.
"""

from __future__ import annotations

import json
import logging
from collections import defaultdict
from dataclasses import dataclass
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    NamedStyle,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models.check_result import CheckResult
from app.models.invest_allocation import (
    BuhCompanyMapping,
    InvestAllocation,
    InvestEmployeeSelection,
    InvestFtePlan,
)
from app.models.upload import Upload
from app.models.worklog import WorklogEntry
from app.services.calendar import get_expected_hours
from app.services.employee_country import get_country
from app.services.invest_summary import (
    aggregate_invest_hours,
    format_project_total_line,
    group_plan_vs_fact_by_project,
    group_saved_allocations,
)
from app.services.permitted_tasks import load_permitted_tasks, sort_invest_projects

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Brandbook tokens → openpyxl primitives
# ---------------------------------------------------------------------------
_CLR_TEXT_PRIMARY = "111827"
_CLR_TEXT_SECONDARY = "4B5563"
_CLR_BG_SECONDARY = "F9FAFB"
_CLR_BORDER = "E5E7EB"

_CLR_SUCCESS = "059669"
_CLR_SUCCESS_LIGHT = "ECFDF5"
_CLR_WARNING = "D97706"
_CLR_WARNING_LIGHT = "FFFBEB"
_CLR_ERROR = "DC2626"
_CLR_ERROR_LIGHT = "FEF2F2"

_THIN_SIDE = Side(style="thin", color=_CLR_BORDER)
_THIN_BORDER = Border(
    left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE
)

_FONT_HEADER = Font(name="Calibri", bold=True, size=14, color=_CLR_TEXT_PRIMARY)
_FONT_SUBHEADER = Font(name="Calibri", bold=True, size=11, color=_CLR_TEXT_PRIMARY)
_FONT_TABLE_HEADER = Font(name="Calibri", bold=True, size=10, color=_CLR_TEXT_SECONDARY)
_FONT_DATA = Font(name="Calibri", size=10, color=_CLR_TEXT_PRIMARY)
_FONT_MONO = Font(name="Consolas", size=10, color=_CLR_TEXT_PRIMARY)
_FONT_SUCCESS = Font(name="Calibri", bold=True, size=10, color=_CLR_SUCCESS)
_FONT_WARNING = Font(name="Calibri", bold=True, size=10, color=_CLR_WARNING)
_FONT_ERROR = Font(name="Calibri", bold=True, size=10, color=_CLR_ERROR)
_FONT_LABEL = Font(name="Calibri", size=10, color=_CLR_TEXT_SECONDARY)

_CLR_ACCENT_LIGHT = "EFF6FF"

_FILL_TABLE_HEADER = PatternFill(start_color=_CLR_BG_SECONDARY, end_color=_CLR_BG_SECONDARY, fill_type="solid")
_FILL_SUCCESS = PatternFill(start_color=_CLR_SUCCESS_LIGHT, end_color=_CLR_SUCCESS_LIGHT, fill_type="solid")
_FILL_WARNING = PatternFill(start_color=_CLR_WARNING_LIGHT, end_color=_CLR_WARNING_LIGHT, fill_type="solid")
_FILL_ERROR = PatternFill(start_color=_CLR_ERROR_LIGHT, end_color=_CLR_ERROR_LIGHT, fill_type="solid")
_FILL_ACCENT = PatternFill(start_color=_CLR_ACCENT_LIGHT, end_color=_CLR_ACCENT_LIGHT, fill_type="solid")

_ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center")

_CHECK_TYPE_LABELS: dict[str, str] = {
    "permitted_task": "Разрешённые задачи",
    "hours_mismatch": "Расхождение часов",
    "comment_quality": "Качество комментариев",
    "comment_relevance": "Релевантность комментариев",
    "time_limit": "Лимит времени",
    "general_rules": "Общие правила",
}

_COUNTRY_LABELS: dict[str, str] = {
    "RU": "Россия",
    "KZ": "Казахстан",
    "BY": "Беларусь",
}

# Downtime task — surfaced as its own report column and sheet.
_DOWNTIME_KEY = "GENERAL-122"
_DOWNTIME_SHEET_TITLE = "Простои GENERAL-122"


# ---------------------------------------------------------------------------
# Named styles (registered once per workbook)
# ---------------------------------------------------------------------------

def _register_styles(wb: Workbook) -> None:
    """Create and register reusable NamedStyle objects."""

    styles: list[NamedStyle] = [
        NamedStyle(
            name="ta_header",
            font=_FONT_HEADER,
            alignment=_ALIGN_LEFT,
        ),
        NamedStyle(
            name="ta_subheader",
            font=_FONT_SUBHEADER,
            alignment=_ALIGN_LEFT,
        ),
        NamedStyle(
            name="ta_table_header",
            font=_FONT_TABLE_HEADER,
            fill=_FILL_TABLE_HEADER,
            border=_THIN_BORDER,
            alignment=_ALIGN_CENTER,
        ),
        NamedStyle(
            name="ta_data",
            font=_FONT_DATA,
            border=_THIN_BORDER,
            alignment=_ALIGN_LEFT,
        ),
        NamedStyle(
            name="ta_mono",
            font=_FONT_MONO,
            border=_THIN_BORDER,
            alignment=_ALIGN_LEFT,
        ),
        NamedStyle(
            name="ta_label",
            font=_FONT_LABEL,
            alignment=_ALIGN_LEFT,
        ),
        NamedStyle(
            name="ta_status_ok",
            font=_FONT_SUCCESS,
            fill=_FILL_SUCCESS,
            border=_THIN_BORDER,
            alignment=_ALIGN_CENTER,
        ),
        NamedStyle(
            name="ta_status_warning",
            font=_FONT_WARNING,
            fill=_FILL_WARNING,
            border=_THIN_BORDER,
            alignment=_ALIGN_CENTER,
        ),
        NamedStyle(
            name="ta_status_error",
            font=_FONT_ERROR,
            fill=_FILL_ERROR,
            border=_THIN_BORDER,
            alignment=_ALIGN_CENTER,
        ),
        NamedStyle(
            name="ta_status_unchecked",
            font=_FONT_LABEL,
            fill=_FILL_TABLE_HEADER,
            border=_THIN_BORDER,
            alignment=_ALIGN_CENTER,
        ),
    ]

    for s in styles:
        wb.add_named_style(s)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _set_column_widths(ws, widths: dict[int, float]) -> None:
    """Set column widths by 1-based column index."""
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _write_row(ws, row: int, values: list, style: str | None = None) -> None:
    """Write a list of values into a row, optionally applying a named style."""
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=val)
        if style:
            cell.style = style


# ---------------------------------------------------------------------------
# Summary sheet builder ("Сводка")
# ---------------------------------------------------------------------------

def _build_summary_sheet(
    ws,
    wl_by_user: dict[str, list[WorklogEntry]],
    cr_by_user: dict[str, list[CheckResult]],
    checks_completed: bool = True,
) -> None:
    """Build the 'Сводка' sheet — one row per employee with hours/status overview."""

    ws.cell(row=1, column=1, value="Сводка по сотрудникам").style = "ta_header"

    headers = [
        "Сотрудник", "Страна", "Факт, ч", "Норма, ч",
        "Разница", "Простои GENERAL-122, ч", "Статус", "Замечания",
    ]
    header_row = 3
    _write_row(ws, header_row, headers, style="ta_table_header")

    total_actual = 0.0
    total_expected = 0.0
    total_downtime = 0.0
    current_row = header_row + 1

    for username in sorted(wl_by_user.keys()):
        user_wl = wl_by_user[username]
        user_cr = cr_by_user.get(username, [])
        country = get_country(username)

        actual = sum(wl.hours for wl in user_wl)
        downtime = sum(wl.hours for wl in user_wl if wl.key == _DOWNTIME_KEY)
        dates = [wl.started for wl in user_wl]
        year = min(dates).year
        month = min(dates).month
        try:
            expected = get_expected_hours(country, year, month)
        except ValueError:
            expected = 0.0
        diff = actual - expected

        has_error = any(cr.severity == "error" for cr in user_cr)
        has_warning = any(cr.severity == "warning" for cr in user_cr)
        if not checks_completed:
            status = "Не проверялось"
        elif has_error:
            status = "Ошибка"
        elif has_warning:
            status = "Внимание"
        else:
            status = "OK"

        issue_count = "—" if not checks_completed else len(user_cr)
        total_actual += actual
        total_expected += expected
        total_downtime += downtime

        row_vals = [
            username,
            country,
            round(actual, 1),
            round(expected, 1),
            round(diff, 1),
            round(downtime, 1),
            status,
            issue_count,
        ]
        _write_row(ws, current_row, row_vals, style="ta_data")

        # Apply monospace to numeric columns
        for col in (3, 4, 5, 6, 8):
            ws.cell(row=current_row, column=col).style = "ta_mono"

        # Highlight downtime hours when present
        if downtime > 0:
            ws.cell(row=current_row, column=6).fill = _FILL_WARNING

        # Status cell coloring
        status_cell = ws.cell(row=current_row, column=7)
        if status == "Ошибка":
            status_cell.style = "ta_status_error"
        elif status == "Внимание":
            status_cell.style = "ta_status_warning"
        elif status == "Не проверялось":
            status_cell.style = "ta_status_unchecked"
        else:
            status_cell.style = "ta_status_ok"

        current_row += 1

    # Totals row
    current_row += 1
    ws.cell(row=current_row, column=1, value="Итого").font = _FONT_SUBHEADER
    ws.cell(row=current_row, column=3, value=round(total_actual, 1)).style = "ta_mono"
    ws.cell(row=current_row, column=3).font = _FONT_SUBHEADER
    ws.cell(row=current_row, column=4, value=round(total_expected, 1)).style = "ta_mono"
    ws.cell(row=current_row, column=4).font = _FONT_SUBHEADER
    ws.cell(row=current_row, column=6, value=round(total_downtime, 1)).style = "ta_mono"
    ws.cell(row=current_row, column=6).font = _FONT_SUBHEADER

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate

    _set_column_widths(ws, {
        1: 30,  # Сотрудник
        2: 14,  # Страна
        3: 12,  # Факт
        4: 12,  # Норма
        5: 12,  # Разница
        6: 22,  # Простои GENERAL-122
        7: 14,  # Статус
        8: 14,  # Замечания
    })


# ---------------------------------------------------------------------------
# Distribution sheet builder ("Распределение")
# ---------------------------------------------------------------------------

def _build_distribution_sheet(
    ws,
    wl_by_user: dict[str, list[WorklogEntry]],
) -> None:
    """Build the 'Распределение' sheet — hours by project × employee pivot."""

    ws.cell(row=1, column=1, value="Распределение часов по проектам").style = "ta_header"

    sorted_users = sorted(wl_by_user.keys())
    projects: set[str] = set()
    hours_matrix: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    for username, entries in wl_by_user.items():
        for wl in entries:
            projects.add(wl.project)
            hours_matrix[wl.project][username] += wl.hours

    sorted_projects = sorted(projects)

    # Column layout: A = Проект, B..N = employees, last = Итого
    header_row = 3
    ws.cell(row=header_row, column=1, value="Проект").style = "ta_table_header"
    for col_idx, uname in enumerate(sorted_users, start=2):
        parts = uname.split()
        short_name = parts[0] if parts else uname
        ws.cell(row=header_row, column=col_idx, value=short_name).style = "ta_table_header"
    total_col = len(sorted_users) + 2
    ws.cell(row=header_row, column=total_col, value="Итого").style = "ta_table_header"

    current_row = header_row + 1
    user_totals: dict[str, float] = defaultdict(float)

    for project in sorted_projects:
        ws.cell(row=current_row, column=1, value=project).style = "ta_data"
        project_total = 0.0
        for col_idx, uname in enumerate(sorted_users, start=2):
            h = hours_matrix[project].get(uname, 0.0)
            val = round(h, 1) if h else None
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.style = "ta_mono"
            project_total += h
            user_totals[uname] += h
        ws.cell(row=current_row, column=total_col, value=round(project_total, 1)).style = "ta_mono"
        current_row += 1

    # "Итого" row per employee
    current_row += 1
    ws.cell(row=current_row, column=1, value="Итого").font = _FONT_SUBHEADER
    grand_total = 0.0
    for col_idx, uname in enumerate(sorted_users, start=2):
        t = user_totals[uname]
        cell = ws.cell(row=current_row, column=col_idx, value=round(t, 1))
        cell.style = "ta_mono"
        cell.font = _FONT_SUBHEADER
        grand_total += t
    cell = ws.cell(row=current_row, column=total_col, value=round(grand_total, 1))
    cell.style = "ta_mono"
    cell.font = _FONT_SUBHEADER

    # --- Region totals section ---
    current_row += 2
    ws.cell(row=current_row, column=1, value="Итого по регионам").style = "ta_subheader"
    current_row += 1

    region_headers = ["Регион", "Сотрудников", "Часов"]
    _write_row(ws, current_row, region_headers, style="ta_table_header")
    current_row += 1

    region_hours: dict[str, float] = defaultdict(float)
    region_count: dict[str, int] = defaultdict(int)
    for uname in sorted_users:
        country = get_country(uname)
        region_hours[country] += user_totals.get(uname, 0.0)
        region_count[country] += 1

    for country in sorted(region_hours.keys()):
        label = _COUNTRY_LABELS.get(country, country)
        _write_row(
            ws, current_row,
            [label, region_count[country], round(region_hours[country], 1)],
            style="ta_data",
        )
        ws.cell(row=current_row, column=2).style = "ta_mono"
        ws.cell(row=current_row, column=3).style = "ta_mono"
        current_row += 1

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate

    _set_column_widths(ws, {1: 24})
    for col_idx in range(2, total_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14


# ---------------------------------------------------------------------------
# Errors sheet builder ("Ошибки")
# ---------------------------------------------------------------------------

_FIX_HINTS: dict[str, str] = {
    "permitted_task": "Перенести списание на разрешённую задачу",
    "comment_quality": "Исправить комментарий согласно требованиям",
    "comment_relevance": "Переписать комментарий, чтобы он соответствовал задаче",
    "time_limit": "Уменьшить время списания (лимит: 20 мин/день)",
    "general_rules": "Перенести списание согласно правилам",
}


def _build_errors_sheet(
    ws,
    check_results: list[CheckResult],
    worklogs: list[WorklogEntry],
) -> None:
    """Build the 'Ошибки' sheet — one row per worklog-level error (excluding hours_mismatch)."""

    ws.cell(row=1, column=1, value="Ошибки в списаниях").style = "ta_header"

    headers = [
        "Key", "Title", "Started", "Username",
        "Time Spent (Hours)", "Comment", "Тип ошибки", "Как исправить",
    ]
    header_row = 3
    _write_row(ws, header_row, headers, style="ta_table_header")

    # Build worklog lookup: (key, username) → WorklogEntry
    wl_lookup: dict[tuple[str, str], WorklogEntry] = {}
    for wl in worklogs:
        wl_lookup.setdefault((wl.key, wl.username), wl)

    current_row = header_row + 1

    for cr in check_results:
        if cr.check_type == "hours_mismatch":
            continue

        check_label = _CHECK_TYPE_LABELS.get(cr.check_type, cr.check_type)
        default_fix = _FIX_HINTS.get(cr.check_type, "")

        try:
            details = json.loads(cr.details)
        except (json.JSONDecodeError, TypeError):
            continue

        if cr.check_type == "permitted_task":
            # details: list of {"key", "hours", "title", "reason", "rule_type"}
            for item in details:
                wl = wl_lookup.get((item["key"], cr.username))
                row_data = [
                    item.get("key", ""),
                    item.get("title", ""),
                    wl.started.strftime("%d.%m.%Y") if wl else "",
                    cr.username,
                    round(item.get("hours", 0), 2),
                    (wl.comment or "") if wl else "",
                    check_label,
                    item.get("reason", default_fix),
                ]
                _write_error_row(ws, current_row, row_data)
                current_row += 1

        elif cr.check_type == "comment_quality":
            # details: list of {"key", "hours", "comment", "severity", "reason"}
            for item in details:
                wl = wl_lookup.get((item["key"], cr.username))
                row_data = [
                    item.get("key", ""),
                    (wl.title if wl else ""),
                    wl.started.strftime("%d.%m.%Y") if wl else "",
                    cr.username,
                    round(item.get("hours", 0), 2),
                    item.get("comment", ""),
                    check_label,
                    item.get("reason", default_fix),
                ]
                _write_error_row(ws, current_row, row_data)
                current_row += 1

        elif cr.check_type == "comment_relevance":
            # details: list of {"key", "hours", "comment", "verdict", "explanation"}
            for item in details:
                wl = wl_lookup.get((item["key"], cr.username))
                row_data = [
                    item.get("key", ""),
                    (wl.title if wl else ""),
                    wl.started.strftime("%d.%m.%Y") if wl else "",
                    cr.username,
                    round(item.get("hours", 0), 2),
                    item.get("comment", ""),
                    check_label,
                    item.get("explanation", default_fix),
                ]
                _write_error_row(ws, current_row, row_data)
                current_row += 1

        elif cr.check_type == "general_rules":
            # details: list of {"key", "hours", "title", "reason"}
            for item in details:
                wl = wl_lookup.get((item["key"], cr.username))
                row_data = [
                    item.get("key", ""),
                    item.get("title", ""),
                    wl.started.strftime("%d.%m.%Y") if wl else "",
                    cr.username,
                    round(item.get("hours", 0), 2),
                    (wl.comment or "") if wl else "",
                    check_label,
                    item.get("reason", default_fix),
                ]
                _write_error_row(ws, current_row, row_data)
                current_row += 1

        elif cr.check_type == "time_limit":
            # details: {"year", "month", "actual_hours", "limit_hours", "working_days", "keys"}
            limited_keys = details.get("keys", [])
            for wl in worklogs:
                if wl.username == cr.username and wl.key in limited_keys:
                    row_data = [
                        wl.key,
                        wl.title,
                        wl.started.strftime("%d.%m.%Y"),
                        cr.username,
                        round(wl.hours, 2),
                        wl.comment or "",
                        check_label,
                        default_fix,
                    ]
                    _write_error_row(ws, current_row, row_data)
                    current_row += 1

    if current_row == header_row + 1:
        ws.cell(row=current_row, column=1, value="Нет ошибок").font = _FONT_SUCCESS
        ws.cell(row=current_row, column=1).fill = _FILL_SUCCESS

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate

    _set_column_widths(ws, {
        1: 16,   # Key
        2: 40,   # Title
        3: 14,   # Started
        4: 28,   # Username
        5: 14,   # Time Spent
        6: 40,   # Comment
        7: 22,   # Тип ошибки
        8: 50,   # Как исправить
    })


def _write_error_row(ws, row: int, values: list) -> None:
    """Write a row into the errors sheet with appropriate styling."""
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        if col_idx == 5:  # hours → monospace
            cell.style = "ta_mono"
        elif col_idx == 1:  # key → monospace
            cell.style = "ta_mono"
        else:
            cell.style = "ta_data"


# ---------------------------------------------------------------------------
# Under-logged hours sheet ("Недобор часов")
# ---------------------------------------------------------------------------

def _build_underlogged_sheet(
    ws,
    check_results: list[CheckResult],
) -> None:
    """Build the 'Недобор часов' sheet — hours_mismatch errors about under-logging."""

    ws.cell(row=1, column=1, value="Недобор часов").style = "ta_header"

    headers = [
        "Сотрудник", "Страна", "Месяц",
        "Факт, ч", "Норма, ч", "Разница, ч",
    ]
    header_row = 3
    _write_row(ws, header_row, headers, style="ta_table_header")

    current_row = header_row + 1
    has_rows = False

    for cr in check_results:
        if cr.check_type != "hours_mismatch":
            continue

        try:
            details = json.loads(cr.details)
        except (json.JSONDecodeError, TypeError):
            continue

        diff = details.get("diff", 0)
        if diff >= 0:
            continue

        has_rows = True
        year = details.get("year", 0)
        month = details.get("month", 0)
        country = details.get("country", "")

        row_data = [
            cr.username,
            country,
            f"{month:02d}.{year}",
            round(details.get("actual_hours", 0), 1),
            round(details.get("expected_hours", 0), 1),
            round(diff, 1),
        ]
        _write_row(ws, current_row, row_data, style="ta_data")

        for col in (4, 5, 6):
            ws.cell(row=current_row, column=col).style = "ta_mono"

        # Highlight the difference cell in red
        diff_cell = ws.cell(row=current_row, column=6)
        diff_cell.font = _FONT_ERROR
        diff_cell.fill = _FILL_ERROR

        current_row += 1

    if not has_rows:
        ws.cell(row=current_row, column=1, value="Нет недобора часов").font = _FONT_SUCCESS
        ws.cell(row=current_row, column=1).fill = _FILL_SUCCESS

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate

    _set_column_widths(ws, {
        1: 30,   # Сотрудник
        2: 14,   # Страна
        3: 12,   # Месяц
        4: 12,   # Факт
        5: 12,   # Норма
        6: 14,   # Разница
    })


# ---------------------------------------------------------------------------
# Downtime sheet builder ("Простои GENERAL-122")
# ---------------------------------------------------------------------------

def _build_downtime_sheet(
    ws,
    worklogs: list[WorklogEntry],
) -> None:
    """Build a dedicated sheet for downtime logged to GENERAL-122."""

    downtime_logs = [wl for wl in worklogs if wl.key == _DOWNTIME_KEY]
    title = (
        downtime_logs[0].title
        if downtime_logs
        else "Простои из-за отсутствия проекта"
    )

    ws.cell(row=1, column=1, value=_DOWNTIME_SHEET_TITLE).style = "ta_header"
    ws.cell(row=2, column=1, value=f"{_DOWNTIME_KEY} — {title}").font = _FONT_LABEL

    # --- Summary by employee ---
    current_row = 4
    ws.cell(row=current_row, column=1, value="Итого по сотрудникам").style = "ta_subheader"
    current_row += 1

    sum_headers = ["Сотрудник", "Страна", "Часы"]
    _write_row(ws, current_row, sum_headers, style="ta_table_header")
    sum_header_row = current_row
    current_row += 1

    by_user: dict[str, float] = defaultdict(float)
    for wl in downtime_logs:
        by_user[wl.username] += wl.hours

    grand_total = 0.0
    if by_user:
        for username in sorted(by_user.keys()):
            hours = by_user[username]
            grand_total += hours
            _write_row(ws, current_row, [
                username,
                get_country(username),
                round(hours, 2),
            ], style="ta_data")
            ws.cell(row=current_row, column=3).style = "ta_mono"
            if hours > 0:
                ws.cell(row=current_row, column=3).fill = _FILL_WARNING
            current_row += 1
    else:
        ws.cell(row=current_row, column=1, value="Нет списаний на простои").font = _FONT_SUCCESS
        ws.cell(row=current_row, column=1).fill = _FILL_SUCCESS
        current_row += 1

    ws.cell(row=current_row, column=1, value="Итого").font = _FONT_SUBHEADER
    cell = ws.cell(row=current_row, column=3, value=round(grand_total, 2))
    cell.style = "ta_mono"
    cell.font = _FONT_SUBHEADER
    current_row += 2

    # --- Detail lines ---
    ws.cell(row=current_row, column=1, value="Детализация списаний").style = "ta_subheader"
    current_row += 1

    detail_headers = ["Сотрудник", "Дата", "Часы", "Комментарий"]
    _write_row(ws, current_row, detail_headers, style="ta_table_header")
    current_row += 1

    if downtime_logs:
        for wl in sorted(downtime_logs, key=lambda w: (w.username, w.started)):
            _write_row(ws, current_row, [
                wl.username,
                wl.started.strftime("%d.%m.%Y") if wl.started else "",
                round(wl.hours, 2),
                wl.comment or "",
            ], style="ta_data")
            ws.cell(row=current_row, column=3).style = "ta_mono"
            current_row += 1
    else:
        ws.cell(row=current_row, column=1, value="Нет записей").font = _FONT_LABEL

    ws.freeze_panes = ws.cell(row=sum_header_row + 1, column=1).coordinate

    _set_column_widths(ws, {
        1: 30,   # Сотрудник
        2: 14,   # Страна / Дата
        3: 12,   # Часы
        4: 60,   # Комментарий
    })


# ---------------------------------------------------------------------------
# Invest directions sheet builder ("Инвест-направления")
# ---------------------------------------------------------------------------

@dataclass
class _AutoRow:
    username: str
    task_key: str
    title: str
    hours: float
    invest_project: str


@dataclass
class _BuhRow:
    username: str
    task_key: str
    title: str
    hours: float
    buh_company: str
    invest_project: str | None
    manual_assigned: bool = False


@dataclass
class _ManualRow:
    username: str
    task_key: str
    title: str
    hours: float
    percentage: float | None
    invest_project: str | None
    allocation_type: str


def _build_invest_sheet(
    ws,
    worklogs: list[WorklogEntry],
    upload_id: int,
    db: Session,
) -> None:
    """Build the 'Инвест-направления' sheet with 4 sections."""

    selected_rows = (
        db.query(InvestEmployeeSelection)
        .filter(InvestEmployeeSelection.upload_id == upload_id)
        .all()
    )
    selected_users = {r.username for r in selected_rows}

    if not selected_users:
        ws.cell(row=1, column=1, value="Инвест-направления").style = "ta_header"
        ws.cell(row=3, column=1, value="Сотрудники не выбраны").font = _FONT_LABEL
        return

    registry = load_permitted_tasks()

    buh_mapping_rows = (
        db.query(BuhCompanyMapping)
        .filter(BuhCompanyMapping.upload_id == upload_id)
        .all()
    )
    buh_map: dict[str, BuhCompanyMapping] = {r.task_key: r for r in buh_mapping_rows}

    saved_alloc_rows = (
        db.query(InvestAllocation)
        .filter(InvestAllocation.upload_id == upload_id)
        .all()
    )
    saved_alloc = group_saved_allocations(saved_alloc_rows)

    fte_plan_rows = (
        db.query(InvestFtePlan)
        .filter(InvestFtePlan.upload_id == upload_id)
        .all()
    )
    fte_by_user: dict[str, dict[str, float]] = defaultdict(dict)
    for r in fte_plan_rows:
        if r.fte_value > 0:
            fte_by_user[r.username][r.invest_project] = r.fte_value

    invest_result = aggregate_invest_hours(
        worklogs=worklogs,
        selected_users=selected_users,
        registry=registry,
        buh_map=buh_map,
        saved_alloc=saved_alloc,
        fte_by_user=fte_by_user,
        auto_agg={},
        buh_agg={},
        manual_agg={},
    )
    auto_rows = invest_result.auto_rows
    buh_rows = invest_result.buh_rows
    manual_rows = invest_result.manual_rows
    plan_vs_fact = invest_result.plan_vs_fact
    project_blocks = group_plan_vs_fact_by_project(plan_vs_fact)

    # --- Compute summary totals per invest project ---
    summary: dict[str, dict[str, float]] = defaultdict(lambda: {
        "auto": 0.0, "buh": 0.0, "manual": 0.0,
    })

    for r in auto_rows:
        summary[r.invest_project]["auto"] += r.hours

    for r in buh_rows:
        if r.invest_project:
            summary[r.invest_project]["buh"] += r.hours

    for r in manual_rows:
        if r.invest_project and r.percentage is not None:
            invest_hours = r.hours * r.percentage / 100.0
            summary[r.invest_project]["manual"] += invest_hours

    # --- Write the sheet ---
    ws.cell(row=1, column=1, value="Инвест-направления").style = "ta_header"
    current_row = 3

    # === SECTION 1: Summary table ===
    ws.cell(row=current_row, column=1, value="Сводка по инвест-проектам").style = "ta_subheader"
    current_row += 1

    sum_headers = [
        "Инвест-проект", "Авто (ч)", "По BUH company (ч)",
        "Ручное распределение (ч)", "Итого (ч)",
    ]
    _write_row(ws, current_row, sum_headers, style="ta_table_header")
    sum_header_row = current_row
    current_row += 1

    grand_auto = grand_buh = grand_manual = grand_total = 0.0
    for proj_name in sort_invest_projects(summary.keys()):
        s = summary[proj_name]
        total = s["auto"] + s["buh"] + s["manual"]
        _write_row(ws, current_row, [
            proj_name,
            round(s["auto"], 2),
            round(s["buh"], 2),
            round(s["manual"], 2),
            round(total, 2),
        ], style="ta_data")
        for col in (2, 3, 4, 5):
            ws.cell(row=current_row, column=col).style = "ta_mono"
        grand_auto += s["auto"]
        grand_buh += s["buh"]
        grand_manual += s["manual"]
        grand_total += total
        current_row += 1

    ws.cell(row=current_row, column=1, value="Итого").font = _FONT_SUBHEADER
    for col, val in [(2, grand_auto), (3, grand_buh), (4, grand_manual), (5, grand_total)]:
        cell = ws.cell(row=current_row, column=col, value=round(val, 2))
        cell.style = "ta_mono"
        cell.font = _FONT_SUBHEADER
    current_row += 2

    # === SECTION 1b: Summary by employee ===
    ws.cell(row=current_row, column=1, value="Итого по сотрудникам").style = "ta_subheader"
    current_row += 1

    emp_sum_headers = [
        "Сотрудник", "Авто (ч)", "По BUH company (ч)",
        "Ручное распределение (ч)", "Итого (ч)",
    ]
    _write_row(ws, current_row, emp_sum_headers, style="ta_table_header")
    current_row += 1

    emp_summary: dict[str, dict[str, float]] = defaultdict(lambda: {
        "auto": 0.0, "buh": 0.0, "manual": 0.0,
    })

    for r in auto_rows:
        emp_summary[r.username]["auto"] += r.hours

    for r in buh_rows:
        if r.invest_project:
            emp_summary[r.username]["buh"] += r.hours

    for r in manual_rows:
        if r.invest_project and r.percentage is not None:
            emp_summary[r.username]["manual"] += r.hours * r.percentage / 100.0

    emp_grand_auto = emp_grand_buh = emp_grand_manual = emp_grand_total = 0.0
    for uname in sorted(emp_summary.keys()):
        s = emp_summary[uname]
        total = s["auto"] + s["buh"] + s["manual"]
        _write_row(ws, current_row, [
            uname,
            round(s["auto"], 2),
            round(s["buh"], 2),
            round(s["manual"], 2),
            round(total, 2),
        ], style="ta_data")
        for col in (2, 3, 4, 5):
            ws.cell(row=current_row, column=col).style = "ta_mono"
        emp_grand_auto += s["auto"]
        emp_grand_buh += s["buh"]
        emp_grand_manual += s["manual"]
        emp_grand_total += total
        current_row += 1

    ws.cell(row=current_row, column=1, value="Итого").font = _FONT_SUBHEADER
    for col, val in [
        (2, emp_grand_auto), (3, emp_grand_buh),
        (4, emp_grand_manual), (5, emp_grand_total),
    ]:
        cell = ws.cell(row=current_row, column=col, value=round(val, 2))
        cell.style = "ta_mono"
        cell.font = _FONT_SUBHEADER
    current_row += 2

    # === SECTION 1c: Plan vs fact grouped by invest project ===
    ws.cell(
        row=current_row, column=1, value="Итоги по проектам"
    ).style = "ta_subheader"
    current_row += 1

    people_headers = [
        "Сотрудник",
        "План, ч",
        "План FTE",
        "Факт, ч",
        "Факт FTE",
        "Разница, ч",
        "Разница FTE",
    ]

    if project_blocks:
        for block in project_blocks:
            title_cell = ws.cell(
                row=current_row, column=1, value=format_project_total_line(block)
            )
            title_cell.font = _FONT_SUBHEADER
            title_cell.alignment = _ALIGN_LEFT
            title_cell.fill = _FILL_ACCENT
            ws.merge_cells(
                start_row=current_row,
                start_column=1,
                end_row=current_row,
                end_column=7,
            )
            for col in range(2, 8):
                ws.cell(row=current_row, column=col).fill = _FILL_ACCENT
                ws.cell(row=current_row, column=col).border = _THIN_BORDER
            title_cell.border = _THIN_BORDER
            if block.delta_fte is not None and abs(block.delta_fte) > 0.05:
                title_cell.fill = _FILL_WARNING
                for col in range(2, 8):
                    ws.cell(row=current_row, column=col).fill = _FILL_WARNING
            current_row += 1

            ws.cell(
                row=current_row, column=1, value="Детализация по людям"
            ).font = _FONT_LABEL
            current_row += 1

            _write_row(ws, current_row, people_headers, style="ta_table_header")
            current_row += 1

            for row in block.people:
                _write_row(ws, current_row, [
                    row.username,
                    row.plan_hours if row.plan_hours is not None else "—",
                    row.plan_fte if row.plan_fte is not None else "—",
                    row.fact_hours,
                    row.fact_fte if row.fact_fte is not None else "—",
                    row.delta_hours if row.delta_hours is not None else "—",
                    row.delta_fte if row.delta_fte is not None else "—",
                ], style="ta_data")
                for col in (2, 3, 4, 5, 6, 7):
                    ws.cell(row=current_row, column=col).style = "ta_mono"
                if row.delta_fte is not None and abs(row.delta_fte) > 0.05:
                    for col in range(1, 8):
                        ws.cell(row=current_row, column=col).fill = _FILL_WARNING
                current_row += 1

            current_row += 1
    else:
        ws.cell(row=current_row, column=1, value="Нет данных").font = _FONT_LABEL
        current_row += 2

    # === SECTION 2: Auto-allocated entries (Type 1) ===
    ws.cell(row=current_row, column=1, value="Автоматическое распределение (100% инвест)").style = "ta_subheader"
    current_row += 1

    auto_headers = ["Сотрудник", "Ключ", "Название задачи", "Часы", "Инвест-проект"]
    _write_row(ws, current_row, auto_headers, style="ta_table_header")
    auto_header_row = current_row
    current_row += 1

    auto_subtotal = 0.0
    if auto_rows:
        for r in auto_rows:
            _write_row(ws, current_row, [
                r.username, r.task_key, r.title,
                round(r.hours, 2), r.invest_project,
            ], style="ta_data")
            ws.cell(row=current_row, column=2).style = "ta_mono"
            ws.cell(row=current_row, column=4).style = "ta_mono"
            auto_subtotal += r.hours
            current_row += 1
    else:
        ws.cell(row=current_row, column=1, value="Нет записей").font = _FONT_LABEL
        current_row += 1

    ws.cell(row=current_row, column=3, value="Итого").font = _FONT_SUBHEADER
    cell = ws.cell(row=current_row, column=4, value=round(auto_subtotal, 2))
    cell.style = "ta_mono"
    cell.font = _FONT_SUBHEADER
    current_row += 2

    # === SECTION 3: BUH company entries (Type 4) ===
    ws.cell(row=current_row, column=1, value="Распределение по BUH Company").style = "ta_subheader"
    current_row += 1

    buh_headers = [
        "Сотрудник", "Ключ", "Название задачи", "Часы",
        "BUH Company", "Инвест-проект",
    ]
    _write_row(ws, current_row, buh_headers, style="ta_table_header")
    buh_header_row = current_row
    current_row += 1

    buh_subtotal = 0.0
    if buh_rows:
        for r in buh_rows:
            proj_display = r.invest_project if r.invest_project else "Не задано"
            _write_row(ws, current_row, [
                r.username, r.task_key, r.title,
                round(r.hours, 2), r.buh_company, proj_display,
            ], style="ta_data")
            ws.cell(row=current_row, column=2).style = "ta_mono"
            ws.cell(row=current_row, column=4).style = "ta_mono"

            if r.invest_project and not r.manual_assigned:
                for col in range(1, 7):
                    ws.cell(row=current_row, column=col).fill = _FILL_SUCCESS
            elif r.invest_project and r.manual_assigned:
                for col in range(1, 7):
                    ws.cell(row=current_row, column=col).fill = _FILL_ACCENT
            else:
                for col in range(1, 7):
                    ws.cell(row=current_row, column=col).fill = _FILL_WARNING

            if r.invest_project:
                buh_subtotal += r.hours
            current_row += 1
    else:
        ws.cell(row=current_row, column=1, value="Нет записей").font = _FONT_LABEL
        current_row += 1

    ws.cell(row=current_row, column=3, value="Итого").font = _FONT_SUBHEADER
    cell = ws.cell(row=current_row, column=4, value=round(buh_subtotal, 2))
    cell.style = "ta_mono"
    cell.font = _FONT_SUBHEADER
    current_row += 2

    # === SECTION 4: Manual allocations (Types 2, 3) ===
    ws.cell(row=current_row, column=1, value="Ручное распределение").style = "ta_subheader"
    current_row += 1

    manual_headers = [
        "Сотрудник", "Ключ", "Название задачи", "Часы",
        "Процент", "Инвест-часы", "Инвест-проект",
    ]
    _write_row(ws, current_row, manual_headers, style="ta_table_header")
    manual_header_row = current_row
    current_row += 1

    manual_subtotal = 0.0
    if manual_rows:
        for r in manual_rows:
            pct = r.percentage if r.percentage is not None else 0.0
            invest_hrs = round(r.hours * pct / 100.0, 2) if r.percentage is not None else 0.0
            proj_display = r.invest_project if r.invest_project else "Не задано"

            _write_row(ws, current_row, [
                r.username, r.task_key, r.title,
                round(r.hours, 2),
                f"{pct:.0f}%" if r.percentage is not None else "—",
                invest_hrs,
                proj_display,
            ], style="ta_data")
            ws.cell(row=current_row, column=2).style = "ta_mono"
            ws.cell(row=current_row, column=4).style = "ta_mono"
            ws.cell(row=current_row, column=5).style = "ta_mono"
            ws.cell(row=current_row, column=6).style = "ta_mono"

            if r.invest_project is None or r.percentage is None:
                for col in range(1, 8):
                    ws.cell(row=current_row, column=col).fill = _FILL_WARNING

            manual_subtotal += invest_hrs
            current_row += 1
    else:
        ws.cell(row=current_row, column=1, value="Нет записей").font = _FONT_LABEL
        current_row += 1

    ws.cell(row=current_row, column=5, value="Итого").font = _FONT_SUBHEADER
    cell = ws.cell(row=current_row, column=6, value=round(manual_subtotal, 2))
    cell.style = "ta_mono"
    cell.font = _FONT_SUBHEADER

    ws.freeze_panes = ws.cell(row=sum_header_row + 1, column=1).coordinate

    _set_column_widths(ws, {
        1: 36,   # Сотрудник / Инвест-проект
        2: 20,   # Ключ / План, ч
        3: 50,   # Название задачи / План FTE
        4: 16,   # Часы / Факт, ч
        5: 26,   # BUH Company / Процент / Факт FTE
        6: 22,   # Инвест-проект / Инвест-часы / Разница, ч
        7: 22,   # Инвест-проект / Разница FTE
    })


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def generate_report(upload_id: int, db: Session) -> bytes:
    """Build an openpyxl Workbook and return .xlsx bytes.

    Sheets: Сводка, Распределение, Инвест-направления, Простои GENERAL-122,
    Ошибки, Недобор часов.
    """

    worklogs: list[WorklogEntry] = (
        db.query(WorklogEntry)
        .filter(WorklogEntry.upload_id == upload_id)
        .all()
    )
    if not worklogs:
        raise ValueError(f"No worklog entries for upload {upload_id}")

    upload = db.query(Upload).filter(Upload.id == upload_id).first()
    checks_completed = upload is not None and upload.status == "checked"

    check_results: list[CheckResult] = (
        db.query(CheckResult)
        .filter(CheckResult.upload_id == upload_id)
        .all()
    )

    # Group by username
    wl_by_user: dict[str, list[WorklogEntry]] = defaultdict(list)
    for wl in worklogs:
        wl_by_user[wl.username].append(wl)

    cr_by_user: dict[str, list[CheckResult]] = defaultdict(list)
    for cr in check_results:
        cr_by_user[cr.username].append(cr)

    wb = Workbook()
    _register_styles(wb)

    # Remove the default sheet created by Workbook()
    default_ws = wb.active
    if default_ws is not None:
        wb.remove(default_ws)

    # 1) Summary sheet — must be first
    ws_summary = wb.create_sheet(title="Сводка")
    _build_summary_sheet(ws_summary, wl_by_user, cr_by_user, checks_completed)

    # 2) Distribution sheet — second
    ws_distribution = wb.create_sheet(title="Распределение")
    _build_distribution_sheet(ws_distribution, wl_by_user)

    # 3) Invest directions sheet — third
    ws_invest = wb.create_sheet(title="Инвест-направления")
    _build_invest_sheet(ws_invest, worklogs, upload_id, db)

    # 4) Downtime (GENERAL-122) sheet
    ws_downtime = wb.create_sheet(title=_DOWNTIME_SHEET_TITLE)
    _build_downtime_sheet(ws_downtime, worklogs)

    # 5) Consolidated errors sheet
    ws_errors = wb.create_sheet(title="Ошибки")
    _build_errors_sheet(ws_errors, check_results, worklogs)

    # 6) Under-logged hours sheet (hours_mismatch with negative diff)
    ws_underlogged = wb.create_sheet(title="Недобор часов")
    _build_underlogged_sheet(ws_underlogged, check_results)

    logger.info(
        "Report generated for upload %d: 6 sheets, %d check results",
        upload_id,
        len(check_results),
    )

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
